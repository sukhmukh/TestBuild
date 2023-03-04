import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium import webdriver
from sample import ExcelUtils
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

file = "pybook.xlsx"
workbook = load_workbook(file)
sheet=workbook.active
row_count = sheet.max_row
col_count = sheet.max_column

for r in range(2,row_count+1):
    driver.get("https://www.amazon.in/")
    driver.maximize_window()
    item= ExcelUtils.readData(file, "Sheet1", r, 1)
    #password = ExcelUtils.readData(file, "Sheet1", r, 2)
    driver.find_element(By.ID, 'twotabsearchtextbox').send_keys(item)
    driver.find_element(By.ID, 'nav-search-submit-button').click()
    time.sleep(6)
    pagetitle=driver.title
    print(pagetitle)
    if item in pagetitle:
        ExcelUtils.writeData(file,"Sheet1",r,3,"Passed")
    else:
        #ExcelUtils.writeData(file,"Sheet1",r,3,pagetitle)
        ExcelUtils.writeData(file,"Sheet1",r,5,"Failed")