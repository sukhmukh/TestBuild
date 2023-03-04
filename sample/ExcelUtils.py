import openpyxl


def readData(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    inputvalue = sheet.cell(row=rownum, column=colnum).value
    return inputvalue


def writeData(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)